# we can import testcases data from excel file instead of python file as an option
import openpyxl
book =openpyxl.load_workbook("D:\\Python_Automation_Udemy\\PythonDemo.xlsx")
sheet =book.active
Dict = {}
cell =sheet.cell(row=1, column=2)
print(cell.value) # read from file
sheet.cell(row=2, column=2).value = "Rahul" # write to the file

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row) # max number of rows in the sheet

print(sheet.max_column) # max number of cols in the sheet

print(sheet['A5'].value) # print certain cell


# following for loop to extract data of testcase2 and put it in dictionary
for i in range(1,sheet.max_row+1):  # to get rows
    if sheet.cell(row =i,column=1).value == "Testcase2":

        for j in range(2,sheet.max_column+1):#to get columns
            #Dict["lastname"]="shetty
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value # lefthand is the key and right hand is the value from the sheet

print(Dict)







